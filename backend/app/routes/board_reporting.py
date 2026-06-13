from __future__ import annotations

from io import BytesIO, StringIO
from datetime import datetime, timedelta, timezone
from typing import Any
import csv
import json
import zipfile

from fastapi import APIRouter, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.authz import require_roles
from app.deps import get_db
from app.db import models

router = APIRouter(tags=["board-reporting"])

GLOBAL_ADMIN_ROLES = {"admin", "super_admin", "security_admin"}


def _user_value(current_user: Any, key: str) -> Any:
    if isinstance(current_user, dict):
        return current_user.get(key)
    return getattr(current_user, key, None)


def _user_email(current_user: Any) -> str:
    return str(_user_value(current_user, "email") or _user_value(current_user, "user_email") or _user_value(current_user, "username") or "").strip().lower()


def _is_global_admin(current_user: Any) -> bool:
    return str(_user_value(current_user, "role") or _user_value(current_user, "role_name") or "") in GLOBAL_ADMIN_ROLES


def _scoped_inspection_rows(db: Session, current_user: Any):
    q = db.query(models.Inspection)
    if _is_global_admin(current_user):
        return q.order_by(models.Inspection.id.desc()).all()
    email = _user_email(current_user)
    return (
        q.join(models.TenantMembership, models.TenantMembership.tenant_id == models.Inspection.tenant_id)
        .filter(models.TenantMembership.user_email == email, models.TenantMembership.is_enabled.is_(True))
        .order_by(models.Inspection.id.desc())
        .all()
    )


def _within_days(dt: datetime | None, days: int) -> bool:
    if not dt:
        return False
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)
    return dt >= now - timedelta(days=days)


def _rows_for_window(db: Session, current_user: Any, days: int):
    rows = _scoped_inspection_rows(db, current_user)
    return [r for r in rows if _within_days(r.created_at, days)]


def _site_name(row: models.Inspection) -> str:
    return (getattr(row, "site_name", None) or "default-site").strip() or "default-site"


def _vendor_name(row: models.Inspection) -> str:
    return (getattr(row, "vendor_name", None) or "unknown").strip() or "unknown"


def _issue_name(row: models.Inspection) -> str:
    return (getattr(row, "detected_issue", None) or "unknown").strip() or "unknown"


def _top_counts(counter: dict[str, int], limit: int = 10):
    return sorted(
        [{"label": k, "count": v} for k, v in counter.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:limit]


def _build_board_report(rows: list[models.Inspection]) -> dict:
    total = len(rows)
    completed = sum(1 for r in rows if (r.status or "").lower() == "completed")
    open_alerts = sum(1 for r in rows if (getattr(r, "alert_status", "open") or "").lower() != "resolved")
    resolved_alerts = sum(1 for r in rows if (getattr(r, "alert_status", "") or "").lower() == "resolved")
    high_risk_count = sum(1 for r in rows if int(getattr(r, "risk_score", 0) or 0) >= 80)
    qa_reviewed = sum(1 for r in rows if (getattr(r, "qa_review_status", "") or "").lower() in {"approved", "overridden"})
    qa_overridden = sum(1 for r in rows if (getattr(r, "qa_review_status", "") or "").lower() == "overridden")

    site_counter: dict[str, int] = {}
    vendor_counter: dict[str, int] = {}
    issue_counter: dict[str, int] = {}

    site_summary: dict[str, dict] = {}
    for r in rows:
        site = _site_name(r)
        vendor = _vendor_name(r)
        issue = _issue_name(r)

        site_counter[site] = site_counter.get(site, 0) + 1
        vendor_counter[vendor] = vendor_counter.get(vendor, 0) + 1
        issue_counter[issue] = issue_counter.get(issue, 0) + 1

        if site not in site_summary:
            site_summary[site] = {
                "site_name": site,
                "total_inspections": 0,
                "open_alerts": 0,
                "resolved_alerts": 0,
                "high_risk_count": 0,
                "qa_reviewed": 0,
                "qa_overridden": 0,
                "avg_confidence_total": 0.0,
            }

        s = site_summary[site]
        s["total_inspections"] += 1
        s["avg_confidence_total"] += float(getattr(r, "confidence", 0.0) or 0.0)

        if (getattr(r, "alert_status", "open") or "").lower() == "resolved":
            s["resolved_alerts"] += 1
        else:
            s["open_alerts"] += 1

        if int(getattr(r, "risk_score", 0) or 0) >= 80:
            s["high_risk_count"] += 1

        if (getattr(r, "qa_review_status", "") or "").lower() in {"approved", "overridden"}:
            s["qa_reviewed"] += 1
        if (getattr(r, "qa_review_status", "") or "").lower() == "overridden":
            s["qa_overridden"] += 1

    site_benchmark = []
    for _, s in site_summary.items():
        reviewed = s["qa_reviewed"] or 1
        total_site = s["total_inspections"] or 1
        site_benchmark.append({
            "site_name": s["site_name"],
            "total_inspections": s["total_inspections"],
            "open_alerts": s["open_alerts"],
            "resolved_alerts": s["resolved_alerts"],
            "high_risk_count": s["high_risk_count"],
            "avg_confidence": round(s["avg_confidence_total"] / total_site, 2),
            "qa_override_rate": round(s["qa_overridden"] / reviewed, 4) if s["qa_reviewed"] else 0.0,
        })

    site_benchmark.sort(key=lambda x: (x["open_alerts"], x["qa_override_rate"], x["total_inspections"]), reverse=True)

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "executive_summary": {
            "total_inspections": total,
            "completed": completed,
            "completion_rate": round(completed / total, 4) if total else 0.0,
            "open_alerts": open_alerts,
            "resolved_alerts": resolved_alerts,
            "high_risk_count": high_risk_count,
            "qa_reviewed": qa_reviewed,
            "qa_overridden": qa_overridden,
            "qa_override_rate": round(qa_overridden / qa_reviewed, 4) if qa_reviewed else 0.0,
            "top_sites": _top_counts(site_counter),
            "top_vendors": _top_counts(vendor_counter),
            "top_issues": _top_counts(issue_counter),
        },
        "site_benchmark": site_benchmark,
        "leadership_narrative": {
            "headline": f"{total} inspections processed in reporting window with {open_alerts} open alerts and {high_risk_count} high-risk findings.",
            "quality_note": f"QA reviewed {qa_reviewed} cases with an override rate of {round((qa_overridden / qa_reviewed) * 100, 1) if qa_reviewed else 0.0}%.",
            "operations_note": f"Top issue trend: {(_top_counts(issue_counter, 1)[0]['label'] if issue_counter else 'none')} | Top site by volume: {(_top_counts(site_counter, 1)[0]['label'] if site_counter else 'none')}.",
        }
    }

    return report


def _csv_text(site_benchmark: list[dict]) -> str:
    output = StringIO()
    if not site_benchmark:
        return ""
    writer = csv.DictWriter(output, fieldnames=list(site_benchmark[0].keys()))
    writer.writeheader()
    writer.writerows(site_benchmark)
    return output.getvalue()


def build_board_report_xlsx_bytes(report: dict) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["metric", "value"])
    for k, v in report["executive_summary"].items():
        if isinstance(v, list):
            ws.append([k, json.dumps(v)])
        else:
            ws.append([k, v])

    ws2 = wb.create_sheet("Leadership Narrative")
    ws2.append(["section", "text"])
    for k, v in report["leadership_narrative"].items():
        ws2.append([k, v])

    ws3 = wb.create_sheet("Site Benchmark")
    items = report["site_benchmark"]
    if items:
        headers = list(items[0].keys())
        ws3.append(headers)
        for item in items:
            ws3.append([item.get(h, "") for h in headers])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@router.get("/board-reporting/weekly")
def board_reporting_weekly(
    days: int = Query(default=7, ge=1, le=30),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _rows_for_window(db, current_user, days)
    report = _build_board_report(rows)
    return JSONResponse(report)


@router.get("/board-reporting/weekly.csv")
def board_reporting_weekly_csv(
    days: int = Query(default=7, ge=1, le=30),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _rows_for_window(db, current_user, days)
    report = _build_board_report(rows)
    text = _csv_text(report["site_benchmark"])
    return StreamingResponse(
        iter([text]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=lumenai_board_ready_weekly.csv"},
    )


@router.get("/board-reporting/weekly.xlsx")
def board_reporting_weekly_xlsx(
    days: int = Query(default=7, ge=1, le=30),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _rows_for_window(db, current_user, days)
    report = _build_board_report(rows)
    content = build_board_report_xlsx_bytes(report)
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lumenai_board_ready_weekly.xlsx"},
    )


@router.get("/board-reporting/weekly.bundle.zip")
def board_reporting_weekly_bundle(
    days: int = Query(default=7, ge=1, le=30),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _rows_for_window(db, current_user, days)
    report = _build_board_report(rows)

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("lumenai_board_ready_weekly.json", json.dumps(report, indent=2))
        zf.writestr("lumenai_board_ready_weekly.csv", _csv_text(report["site_benchmark"]))
        zf.writestr("lumenai_board_ready_weekly.xlsx", build_board_report_xlsx_bytes(report))
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=lumenai_board_ready_weekly_bundle.zip"},
    )
