from __future__ import annotations

from io import StringIO, BytesIO
from typing import Any
import csv
import json
import zipfile

from fastapi import APIRouter, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.authz import require_roles
from app.deps import get_db
from app.db import models

router = APIRouter(tags=["review-analytics"])


GLOBAL_ADMIN_ROLES = {"admin", "super_admin", "security_admin"}


def _user_value(current_user: Any, key: str) -> Any:
    if isinstance(current_user, dict):
        return current_user.get(key)
    return getattr(current_user, key, None)


def _user_email(current_user: Any) -> str:
    return str(_user_value(current_user, "email") or _user_value(current_user, "user_email") or _user_value(current_user, "username") or "").strip().lower()


def _is_global_admin(current_user: Any) -> bool:
    return str(_user_value(current_user, "role") or _user_value(current_user, "role_name") or "") in GLOBAL_ADMIN_ROLES


def _scoped_inspection_rows(db: Session, current_user: Any):
    q = db.query(models.Inspection)
    if _is_global_admin(current_user):
        return q.order_by(models.Inspection.id.desc()).all()
    email = _user_email(current_user)
    return (
        q.join(models.TenantMembership, models.TenantMembership.tenant_id == models.Inspection.tenant_id)
        .filter(models.TenantMembership.user_email == email, models.TenantMembership.is_enabled.is_(True))
        .order_by(models.Inspection.id.desc())
        .all()
    )


def _inspection_feedback_row(r: models.Inspection) -> dict:
    return {
        "inspection_id": r.id,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "file_name": r.file_name,
        "vendor_name": r.vendor_name,
        "model_name": r.model_name,
        "model_version": r.model_version,
        "confidence": r.confidence,
        "original_stain_detected": r.stain_detected,
        "original_material_type": r.material_type,
        "original_instrument_type": r.instrument_type,
        "original_detected_issue": r.detected_issue,
        "original_risk_score": r.risk_score,
        "qa_review_status": r.qa_review_status,
        "qa_reviewer": r.qa_reviewer,
        "qa_review_notes": r.qa_review_notes,
        "qa_reviewed_at": r.qa_reviewed_at.isoformat() if r.qa_reviewed_at else None,
        "qa_override_stain_detected": r.qa_override_stain_detected,
        "qa_override_material_type": r.qa_override_material_type,
        "qa_override_instrument_type": r.qa_override_instrument_type,
        "qa_override_detected_issue": r.qa_override_detected_issue,
        "qa_override_risk_score": r.qa_override_risk_score,
        "final_stain_detected": r.qa_override_stain_detected if r.qa_override_stain_detected is not None else r.stain_detected,
        "final_material_type": r.qa_override_material_type or r.material_type,
        "final_instrument_type": r.qa_override_instrument_type or r.instrument_type,
        "final_detected_issue": r.qa_override_detected_issue or r.detected_issue,
        "final_risk_score": r.qa_override_risk_score if r.qa_override_risk_score is not None else r.risk_score,
    }


def _is_override(r: models.Inspection) -> bool:
    return (r.qa_review_status or "").lower() == "overridden"


def _review_summary(rows: list[models.Inspection]) -> dict:
    total_reviewed = sum(1 for r in rows if (r.qa_review_status or "").lower() in {"approved", "overridden"})
    total_pending = sum(1 for r in rows if (r.qa_review_status or "").lower() == "pending")
    total_approved = sum(1 for r in rows if (r.qa_review_status or "").lower() == "approved")
    total_overridden = sum(1 for r in rows if _is_override(r))

    override_by_issue: dict[str, int] = {}
    override_by_vendor: dict[str, int] = {}
    reviewer_counts: dict[str, int] = {}

    for r in rows:
        reviewer = (r.qa_reviewer or "").strip()
        if reviewer:
            reviewer_counts[reviewer] = reviewer_counts.get(reviewer, 0) + 1

        if _is_override(r):
            issue = (r.detected_issue or "unknown").strip() or "unknown"
            vendor = (r.vendor_name or "unknown").strip() or "unknown"
            override_by_issue[issue] = override_by_issue.get(issue, 0) + 1
            override_by_vendor[vendor] = override_by_vendor.get(vendor, 0) + 1

    def top_items(d: dict[str, int], limit: int = 10):
        return sorted(
            [{"label": k, "count": v} for k, v in d.items()],
            key=lambda x: x["count"],
            reverse=True,
        )[:limit]

    agreement_rate = round((total_approved / total_reviewed), 4) if total_reviewed else 0.0
    override_rate = round((total_overridden / total_reviewed), 4) if total_reviewed else 0.0

    return {
        "total_reviewed": total_reviewed,
        "total_pending": total_pending,
        "total_approved": total_approved,
        "total_overridden": total_overridden,
        "agreement_rate": agreement_rate,
        "override_rate": override_rate,
        "top_override_issues": top_items(override_by_issue),
        "top_override_vendors": top_items(override_by_vendor),
        "top_reviewers": top_items(reviewer_counts),
    }


def _csv_text(items: list[dict]) -> str:
    output = StringIO()
    if not items:
        output.write("")
        return output.getvalue()

    writer = csv.DictWriter(output, fieldnames=list(items[0].keys()))
    writer.writeheader()
    writer.writerows(items)
    return output.getvalue()


def _xlsx_bytes(summary: dict, items: list[dict]) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Review Summary"
    ws1.append(["metric", "value"])
    ws1.append(["total_reviewed", summary["total_reviewed"]])
    ws1.append(["total_pending", summary["total_pending"]])
    ws1.append(["total_approved", summary["total_approved"]])
    ws1.append(["total_overridden", summary["total_overridden"]])
    ws1.append(["agreement_rate", summary["agreement_rate"]])
    ws1.append(["override_rate", summary["override_rate"]])

    ws2 = wb.create_sheet("Feedback Dataset")
    if items:
      ws2.append(list(items[0].keys()))
      for item in items:
          ws2.append([item.get(k, "") for k in items[0].keys()])

    ws3 = wb.create_sheet("Override Issues")
    ws3.append(["issue", "count"])
    for item in summary["top_override_issues"]:
        ws3.append([item["label"], item["count"]])

    ws4 = wb.create_sheet("Override Vendors")
    ws4.append(["vendor", "count"])
    for item in summary["top_override_vendors"]:
        ws4.append([item["label"], item["count"]])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@router.get("/review-analytics/summary")
def review_analytics_summary(
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _scoped_inspection_rows(db, current_user)
    return JSONResponse(_review_summary(rows))


@router.get("/review-analytics/feedback-dataset.json")
def feedback_dataset_json(
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _scoped_inspection_rows(db, current_user)
    items = [_inspection_feedback_row(r) for r in rows if (r.qa_review_status or "").lower() in {"approved", "overridden"}]
    return JSONResponse({"items": items})


@router.get("/review-analytics/feedback-dataset.csv")
def feedback_dataset_csv(
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _scoped_inspection_rows(db, current_user)
    items = [_inspection_feedback_row(r) for r in rows if (r.qa_review_status or "").lower() in {"approved", "overridden"}]
    text = _csv_text(items)
    return StreamingResponse(
        iter([text]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=lumenai_feedback_dataset.csv"},
    )


@router.get("/review-analytics/feedback-dataset.xlsx")
def feedback_dataset_xlsx(
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _scoped_inspection_rows(db, current_user)
    summary = _review_summary(rows)
    items = [_inspection_feedback_row(r) for r in rows if (r.qa_review_status or "").lower() in {"approved", "overridden"}]
    content = _xlsx_bytes(summary, items)
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lumenai_feedback_dataset.xlsx"},
    )


@router.get("/review-analytics/feedback-dataset.bundle.zip")
def feedback_dataset_bundle(
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "spd_manager")),
):
    rows = _scoped_inspection_rows(db, current_user)
    summary = _review_summary(rows)
    items = [_inspection_feedback_row(r) for r in rows if (r.qa_review_status or "").lower() in {"approved", "overridden"}]

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("lumenai_review_summary.json", json.dumps(summary, indent=2))
        zf.writestr("lumenai_feedback_dataset.json", json.dumps({"items": items}, indent=2))
        zf.writestr("lumenai_feedback_dataset.csv", _csv_text(items))
        zf.writestr("lumenai_feedback_dataset.xlsx", _xlsx_bytes(summary, items))
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=lumenai_feedback_dataset_bundle.zip"},
    )
